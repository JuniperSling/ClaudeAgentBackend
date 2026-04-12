"""Download and extract text from files sent via QQ."""

import logging
import os
import re
from pathlib import Path

import httpx

logger = logging.getLogger(__name__)

EXTRACTABLE_EXTENSIONS = {
    ".txt", ".md", ".csv", ".json", ".xml", ".yaml", ".yml",
    ".py", ".js", ".ts", ".go", ".java", ".c", ".cpp", ".h",
    ".html", ".css", ".sql", ".sh", ".bat", ".log", ".ini", ".conf",
    ".docx", ".xlsx", ".pdf",
}

MAX_EXTRACT_SIZE = 15000


async def download_file(
    http_client: httpx.AsyncClient,
    file_id: str,
    save_dir: str,
    direct_url: str = "",
    file_name_hint: str = "",
    group_id: str | None = None,
) -> dict | None:
    """Download a file. Uses get_group_file_url/get_private_file_url to get URL, then downloads."""
    os.makedirs(save_dir, exist_ok=True)
    file_name = file_name_hint or "unknown"
    target_path = os.path.join(save_dir, file_name)

    # Priority 1: get_file API → check shared volume for local file (fastest)
    if file_id:
        try:
            logger.info("Calling get_file: file_id=%s", file_id[:40])
            resp = await http_client.post("/get_file", json={"file_id": file_id}, timeout=300)
            resp.raise_for_status()
            data = resp.json()
            if data.get("status") == "ok" and data.get("data"):
                file_info = data["data"]

                napcat_path = file_info.get("file", "")
                if napcat_path and napcat_path.startswith("/app/.config/QQ/"):
                    shared_path = napcat_path.replace("/app/.config/QQ/", "/napcat_files/", 1)
                    if os.path.exists(shared_path):
                        import shutil
                        shutil.copy2(shared_path, target_path)
                        logger.info("File saved (shared volume): %s (%s bytes)", target_path, os.path.getsize(target_path))
                        return {"file_name": file_name, "local_path": target_path, "file_size": str(os.path.getsize(target_path))}

                b64_data = file_info.get("base64", "")
                if b64_data:
                    import base64
                    with open(target_path, "wb") as f:
                        f.write(base64.b64decode(b64_data))
                    logger.info("File saved (base64): %s (%s bytes)", target_path, os.path.getsize(target_path))
                    return {"file_name": file_name, "local_path": target_path, "file_size": str(os.path.getsize(target_path))}
            else:
                logger.warning("get_file returned: %s", data.get("message", "unknown"))
        except Exception:
            logger.exception("get_file failed for %s", file_id)

    # Priority 2: get_group_file_url / get_private_file_url → stream download
    download_url = ""
    if file_id:
        try:
            if group_id:
                logger.info("Calling get_group_file_url: file_id=%s, group=%s", file_id[:40], group_id)
                resp = await http_client.post(
                    "/get_group_file_url",
                    json={"file_id": file_id, "group_id": int(group_id)},
                    timeout=30,
                )
            else:
                logger.info("Calling get_private_file_url: file_id=%s", file_id[:40])
                resp = await http_client.post(
                    "/get_private_file_url",
                    json={"file_id": file_id},
                    timeout=30,
                )
            resp.raise_for_status()
            data = resp.json()
            if data.get("status") == "ok" and data.get("data"):
                download_url = data["data"].get("url", "")
                logger.info("Got file URL: %s", download_url[:100] if download_url else "empty")
            else:
                logger.warning("get_file_url failed: %s", data.get("message", data))
        except Exception:
            logger.exception("Failed to get file URL for %s", file_id)

    if download_url:
        try:
            logger.info("Downloading file: %s -> %s", download_url[:80], target_path)
            async with httpx.AsyncClient(timeout=300, follow_redirects=True) as dl:
                async with dl.stream("GET", download_url) as stream:
                    stream.raise_for_status()
                    with open(target_path, "wb") as f:
                        async for chunk in stream.aiter_bytes(8192):
                            f.write(chunk)
            size = os.path.getsize(target_path)
            logger.info("File saved: %s (%s bytes)", target_path, size)
            return {"file_name": file_name, "local_path": target_path, "file_size": str(size)}
        except Exception:
            logger.exception("File download failed: %s", download_url[:80])

    logger.warning("All download methods failed for file: %s", file_name)
    return None


def extract_text(file_path: str) -> str | None:
    """Extract readable text from a file. Returns None if not extractable."""
    ext = Path(file_path).suffix.lower()
    if ext not in EXTRACTABLE_EXTENSIONS:
        return None

    try:
        if ext == ".docx":
            return _extract_docx(file_path)
        elif ext == ".xlsx":
            return _extract_xlsx(file_path)
        elif ext == ".pdf":
            return _extract_pdf(file_path)
        else:
            with open(file_path, "r", errors="replace") as f:
                text = f.read(MAX_EXTRACT_SIZE)
            if len(text) >= MAX_EXTRACT_SIZE:
                text += "\n\n...(文件内容过长，已截断)"
            return text
    except Exception:
        logger.exception("Failed to extract text from %s", file_path)
        return None


def _extract_docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    text = "\n".join(paragraphs)
    if len(text) > MAX_EXTRACT_SIZE:
        text = text[:MAX_EXTRACT_SIZE] + "\n\n...(文件内容过长，已截断)"
    return text


def _extract_xlsx(path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    lines = []
    for sheet in wb.sheetnames[:3]:
        ws = wb[sheet]
        lines.append(f"=== Sheet: {sheet} ===")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 50:
                lines.append("...(行数过多，已截断)")
                break
            lines.append("\t".join(str(c) if c is not None else "" for c in row))
    text = "\n".join(lines)
    if len(text) > MAX_EXTRACT_SIZE:
        text = text[:MAX_EXTRACT_SIZE] + "\n\n...(文件内容过长，已截断)"
    return text


def _extract_pdf(path: str) -> str:
    from PyPDF2 import PdfReader
    reader = PdfReader(path)
    pages = []
    for i, page in enumerate(reader.pages[:20]):
        text = page.extract_text()
        if text:
            pages.append(f"--- Page {i+1} ---\n{text}")
    text = "\n\n".join(pages)
    if len(text) > MAX_EXTRACT_SIZE:
        text = text[:MAX_EXTRACT_SIZE] + "\n\n...(文件内容过长，已截断)"
    return text
